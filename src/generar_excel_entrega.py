from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


PROJECT_ROOT = Path(".")
TEST_XLSX = PROJECT_ROOT / "dataInicial" / "dataset_credito-test.xlsx"
SUBMISSION_CSV = PROJECT_ROOT / "submission.csv"
OUTPUT_XLSX = PROJECT_ROOT / "GRUPOArmonitech.xlsx"


def main() -> None:
    submission = pd.read_csv(SUBMISSION_CSV)
    required = {"id_cliente", "prob_default"}
    missing = required - set(submission.columns)
    if missing:
        raise ValueError(f"Faltan columnas en submission.csv: {sorted(missing)}")
    if not submission["prob_default"].between(0, 1).all():
        raise ValueError("prob_default contiene valores fuera de [0, 1]")
    if not submission["id_cliente"].is_unique:
        raise ValueError("submission.csv contiene id_cliente duplicados")

    proba_by_id = dict(zip(submission["id_cliente"], submission["prob_default"]))

    wb = load_workbook(TEST_XLSX)
    ws = wb.active

    headers = {cell.value: cell.column for cell in ws[1]}
    if "id_cliente" not in headers:
        raise ValueError("No existe columna id_cliente en el Excel test")
    if "Probabilidad" not in headers:
        raise ValueError("No existe columna Probabilidad en el Excel test")

    id_col = headers["id_cliente"]
    prob_col = headers["Probabilidad"]
    yellow = PatternFill(fill_type="solid", fgColor="FFFF00")

    ws.cell(row=1, column=prob_col).fill = yellow
    matched = 0
    missing_ids = []
    for row in range(2, ws.max_row + 1):
        id_value = ws.cell(row=row, column=id_col).value
        if id_value not in proba_by_id:
            missing_ids.append(id_value)
            continue
        cell = ws.cell(row=row, column=prob_col)
        cell.value = float(proba_by_id[id_value])
        cell.number_format = "0.000000"
        cell.fill = yellow
        matched += 1

    if missing_ids:
        raise ValueError(f"IDs del test sin probabilidad: {missing_ids[:10]}")
    if matched != len(submission):
        raise ValueError(f"Filas matcheadas={matched}, predicciones={len(submission)}")

    ws.column_dimensions[ws.cell(row=1, column=prob_col).column_letter].width = 14
    wb.save(OUTPUT_XLSX)
    print(f"Excel final generado: {OUTPUT_XLSX}")
    print(f"Filas con Probabilidad llenada: {matched}")


if __name__ == "__main__":
    main()
